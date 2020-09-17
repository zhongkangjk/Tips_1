import os,sys

def 返回文件绝对路径(当前路径文件名):
    if hasattr(sys, '_MEIPASS'):
        # PyInstaller会创建临时文件夹temp
        # 并把路径存储在_MEIPASS中
        base_path = sys._MEIPASS
    else:
        base_path = os.path.abspath('.')
    return os.path.join(base_path, 当前路径文件名)

#print(返回文件绝对路径('list1.py'))


import os

def 返回当前路径第一个包含名字的文件(名字):
    namelist = os.listdir(os.getcwd())
    for x in namelist:
        if 名字 in x and '$' not in x:
            路径 = x
            return 路径
            break
#print(返回当前路径第一个包含名字的文件('lis'))


import xlrd

class 打开excel文件():
    def __init__(self,文件名,第几个表):
        self.文件 = xlrd.open_workbook(filename = 文件名)
        self.表 = self.文件.sheet_by_index(第几个表-1)


    def 获得横向资料(self):
        self.数据 = [self.表.row_values(i) for i in range(self.表.nrows)]
        return self.数据


    def 获得纵向资料(self):
        self.数据 =  [self.表.col_values(i) for i in range(self.表.ncols)]
        return self.数据

    def 获得名称列数据(self,名称):
        def 获得列序号(表名,查找字段名):
            列序号 = None
            for i in range(表名.ncols):
                if (表名.cell_value(0,i) == 查找字段名):
                    列序号 = i
                    break
            return 列序号
        print(获得列序号(self.表,名称))
        self.数据 = self.表.col_values(获得列序号(self.表,名称),2)
        return self.数据

# 文件 = 打开excel文件('测试.xlsx',1)
# print(文件.获得名称列数据('地址'))


import re

def 列表取第一个电话(list1):
    list2 = []
    for i in list1:
        for j in i:
            if re.match(r"^1\d{10}$", j):
                list2.append(j)
                break
        else:
            list2.append('mei')
    return list2
#list1 = [['123','13210000000','13210000000'],['159','15988886666'],['123']]
#print(列表取第一个电话(list1))


def 求列表元素出现次数字典(list):
    结果 = {}
    for i in set(list):
        结果[i] = list.count(i)
    return 结果


import pandas as pd

class 提取类():
    def __init__(self,文件名,第几个表):
        self.文件 = pd.read_excel(文件名, 第几个表)

    def 拿取数据(self,从第几行开始,*列名):
        self.数据 = self.文件.loc[从第几行开始-2:,[*列名,]]
        #print(self.数据)
        return self.数据
# 数据文件名 = "名称.xlsx"
# def 销售表提取():
#     销售表 = 提取类(数据文件名,0)
#     销售数据 = 销售表.拿取数据(8234,'企业名称','日期','服务费发票','金额','支付方式')
#     销售数据['发票类型'] = '纸票'
#     销售数据['日期'].fillna(method="ffill",inplace=True) #空着的日期按照前一个来
#     销售数据['日期'] = 销售数据['日期'].apply(lambda x:x.strftime('%Y-%m-%d') if x == x and type(x) != type('sd') else '日期为空')
#     销售数据 = 销售数据.values.tolist()


import openpyxl
from openpyxl.styles import Alignment, Font

def openpyxl写新文件():
    填充的数据列表 = []

    结果文件 = openpyxl.Workbook()

    表1 = 结果文件.active

    表1.title = '表1'
    表1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    表1.cell(1, 1).value = '这里我就写个字'

    标题 = 表1['A1']
    标题.font = Font(name = '黑体',size = 20)
    标题.alignment = Alignment(horizontal='center', vertical='center')

    表1.column_dimensions['B'].width = 33
    表1.column_dimensions['C'].width = 11

    for i in 填充的数据列表:
        表1.append(i)

    表2 = 结果文件.create_sheet('表2')
    表2.append('数据')

    结果文件.save('结果文件.xlsx')


